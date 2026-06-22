from dotenv import load_dotenv
load_dotenv(override=False)
from flask import Flask, render_template, request, redirect, url_for, session
from datetime import timedelta, datetime
import mysql.connector
import os

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'troque-esta-chave-antes-de-usar')
app.permanent_session_lifetime = timedelta(minutes=30)

def formatar_moeda(valor):
    if valor is None:
        return 'R$ 0,00'
    return f"R$ {float(valor):_.2f}".replace('.', ',').replace('_', '.')

def formatar_documento(doc):
    if not doc:
        return ''
    d = ''.join(filter(str.isdigit, str(doc)))
    if len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    elif len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return doc

app.jinja_env.globals['formatar_moeda']     = formatar_moeda
app.jinja_env.globals['formatar_documento'] = formatar_documento

USUARIOS = {
    'Admin': 'ecil2026'
}

def conectar():
    return mysql.connector.connect(
        host=os.environ.get('DB_HOST', 'localhost'),
        port=int(os.environ.get('DB_PORT', 3306)),
        database=os.environ.get('DB_NAME', 'analisecredito_db'),
        user=os.environ.get('DB_USER', 'root'),
        password=os.environ.get('DB_PASSWORD', '')
    )

def login_necessario():
    return 'usuario' not in session

def limpar_valor(v, padrao=0):
    try:
        return float(str(v).replace('.', '').replace(',', '.'))
    except:
        return padrao

# ─── LOGIN / LOGOUT ───────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        usuario = request.form['usuario']
        senha   = request.form['senha']
        if usuario in USUARIOS and USUARIOS[usuario] == senha:
            session.permanent = True
            session['usuario'] = usuario
            return redirect(url_for('index'))
        erro = '❌ Usuário ou senha incorretos!'
    return render_template('login.html', erro=erro)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── PÁGINA INICIAL ───────────────────────────────────────────

@app.route('/')
def index():
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT C.ClienteID, C.Nome, C.CPF, C.Situacao,
           (SELECT AC.PontuacaoCredito FROM AnalisesCredito AC
            WHERE AC.ClienteID = C.ClienteID
            ORDER BY AC.DataAnalise DESC LIMIT 1) AS Score
        FROM Clientes C
        ORDER BY C.Nome
    """)
    clientes = cursor.fetchall()

    aprovados  = sum(1 for c in clientes if c[3] == 'Aprovado')
    pendentes  = sum(1 for c in clientes if c[3] == 'Pendente')
    recusados  = sum(1 for c in clientes if c[3] == 'Recusado')

    score_excelente = sum(1 for c in clientes if c[4] and int(c[4]) >= 851)
    score_bom       = sum(1 for c in clientes if c[4] and 701 <= int(c[4]) <= 850)
    score_regular   = sum(1 for c in clientes if c[4] and 501 <= int(c[4]) <= 700)
    score_baixo     = sum(1 for c in clientes if c[4] and int(c[4]) <= 500)

    cursor.execute("SELECT COUNT(*) FROM Fornecedores")
    total_fornecedores = cursor.fetchone()[0]

    conn.close()
    return render_template('index.html',
        clientes=clientes,
        aprovados=aprovados,
        pendentes=pendentes,
        recusados=recusados,
        score_excelente=score_excelente,
        score_bom=score_bom,
        score_regular=score_regular,
        score_baixo=score_baixo,
        total_fornecedores=total_fornecedores)

# ─── CLIENTES ─────────────────────────────────────────────────

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    if login_necessario():
        return redirect(url_for('login'))
    mensagem = None
    if request.method == 'POST':
        nome      = request.form.get('nome', '')
        cpf       = request.form.get('cpf', '')
        historico = request.form.get('historico', 'Bom')
        dividas   = limpar_valor(request.form.get('dividas', 0))
        setor     = request.form.get('setor', 'Pescados')
        serasa    = request.form.get('serasa', '')
        score_raw = request.form.get('score', '0')
        conn   = conectar()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO Clientes
            (Nome, CPF, RendaMensal, HistoricoPagamento, DividasAtuais, SetorAtuacao, SerasaObservacao)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (nome, cpf, 0, historico, dividas, setor, serasa))
        cliente_id = cursor.lastrowid
        if score_raw and int(score_raw) > 0:
            score = int(score_raw)
            if score >= 851:
                taxa = 1.50
            elif score >= 701:
                taxa = 3.00
            elif score >= 501:
                taxa = 5.00
            elif score >= 301:
                taxa = 8.00
            else:
                taxa = 12.00

            decisao = 'Pendente'
            limite = limpar_valor(request.form.get('limite', 0))
            cursor.execute("""
                INSERT INTO AnalisesCredito
                (ClienteID, PontuacaoCredito, DecisaoCredito, LimiteSugerido, TaxaJuros)
                VALUES (%s, %s, %s, %s, %s)
            """, (cliente_id, score, decisao, limite, taxa))
            cursor.execute("""
                UPDATE Clientes SET Situacao=%s WHERE ClienteID=%s
            """, ('Pendente', cliente_id))
        conn.commit()
        conn.close()
        mensagem = f'✅ Cliente "{nome}" cadastrado com sucesso!'
    return render_template('cadastro.html', mensagem=mensagem)

@app.route('/editar_cliente/<int:id>', methods=['GET', 'POST'])
def editar_cliente(id):
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    if request.method == 'POST':
        nome      = request.form.get('nome', '')
        cpf       = request.form.get('cpf', '')
        historico = request.form.get('historico', 'Bom')
        dividas   = limpar_valor(request.form.get('dividas', 0))
        setor     = request.form.get('setor', 'Pescados')
        situacao  = request.form.get('situacao', 'Pendente')
        serasa    = request.form.get('serasa', '')
        limite    = limpar_valor(request.form.get('limite', 0))

        if limite > 0:
            cursor.execute("""
                UPDATE AnalisesCredito SET LimiteSugerido=%s
                WHERE ClienteID=%s
            """, (limite, id))

        cursor.execute("""
            UPDATE Clientes SET
                Nome=%s, CPF=%s, HistoricoPagamento=%s,
                DividasAtuais=%s, SetorAtuacao=%s, Situacao=%s, SerasaObservacao=%s
            WHERE ClienteID=%s
        """, (nome, cpf, historico, dividas, setor, situacao, serasa, id))
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    cursor.execute("""
        SELECT ClienteID, Nome, CPF, RendaMensal, HistoricoPagamento,
               DividasAtuais, SetorAtuacao, Situacao, SerasaObservacao
        FROM Clientes WHERE ClienteID=%s
    """, (id,))
    cliente = cursor.fetchone()
    conn.close()
    return render_template('editar_cliente.html', cliente=cliente)

@app.route('/excluir_cliente/<int:id>')
def excluir_cliente(id):
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Clientes WHERE ClienteID=%s", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/situacao/<int:id>/<string:situacao>')
def alterar_situacao(id, situacao):
    if login_necessario():
        return redirect(url_for('login'))
    if situacao not in ['Aprovado', 'Recusado', 'Pendente']:
        return redirect(url_for('index'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("UPDATE Clientes SET Situacao=%s WHERE ClienteID=%s", (situacao, id))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

# ─── ANÁLISE ──────────────────────────────────────────────────

@app.route('/analise')
def analise():
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT C.Nome, C.CPF, AC.PontuacaoCredito, C.Situacao,
               AC.LimiteSugerido, AC.TaxaJuros, AC.DataAnalise
        FROM AnalisesCredito AC
        INNER JOIN Clientes C ON AC.ClienteID = C.ClienteID
        ORDER BY AC.DataAnalise DESC
    """)
    analises = cursor.fetchall()
    conn.close()
    return render_template('analise.html', analises=analises)

# ─── FORNECEDORES ─────────────────────────────────────────────

@app.route('/fornecedores')
def fornecedores():
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT C.Nome, C.CPF, F.NomeFornecedor, F.Email,
               F.Telefone, F.NomeContato, F.ForneceReferencia,
               F.LimiteCredito, F.UltimaCompraData, F.FornecedorID
        FROM Fornecedores F
        INNER JOIN Clientes C ON F.ClienteID = C.ClienteID
        ORDER BY C.Nome, F.NomeFornecedor
    """)
    fornecedores = cursor.fetchall()
    conn.close()
    return render_template('fornecedores.html', fornecedores=fornecedores)

@app.route('/cadastro_fornecedor', methods=['GET', 'POST'])
def cadastro_fornecedor():
    if login_necessario():
        return redirect(url_for('login'))
    mensagem = None
    erro     = None
    conn     = conectar()
    cursor   = conn.cursor()
    cursor.execute("SELECT ClienteID, Nome, CPF FROM Clientes ORDER BY Nome")
    clientes = cursor.fetchall()
    if request.method == 'POST':
        try:
            cliente_id          = request.form.get('cliente_id')
            nome_fornecedor     = request.form.get('nome_fornecedor', '')
            email               = request.form.get('email', '')
            telefone            = request.form.get('telefone', '')
            nome_contato        = request.form.get('nome_contato', '')
            fornece_referencia  = request.form.get('fornece_referencia', 'Sim')
            limite_credito      = limpar_valor(request.form.get('limite_credito', 0))
            maior_compra_data   = request.form.get('maior_compra_data') or None
            maior_compra_valor  = limpar_valor(request.form.get('maior_compra_valor', 0))
            ultima_compra_data  = request.form.get('ultima_compra_data') or None
            ultima_compra_valor = limpar_valor(request.form.get('ultima_compra_valor', 0))
            data_cadastro       = request.form.get('data_cadastro_fornecedor') or None
            cursor.execute("""
                INSERT INTO Fornecedores
                (ClienteID, NomeFornecedor, Email, Telefone, NomeContato,
                 ForneceReferencia, LimiteCredito, PrimeiraCompraData,
                 PrimeiraCompraValor, UltimaCompraData, UltimaCompraValor,
                 DataCadastroFornecedor)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (cliente_id, nome_fornecedor, email, telefone, nome_contato,
                  fornece_referencia, limite_credito, maior_compra_data,
                  maior_compra_valor, ultima_compra_data, ultima_compra_valor,
                  data_cadastro))
            conn.commit()
            mensagem = f'✅ Fornecedor "{nome_fornecedor}" cadastrado com sucesso!'
        except Exception as e:
            erro = f'❌ Erro ao cadastrar: {str(e)}'
    conn.close()
    return render_template('cadastro_fornecedor.html',
                           clientes=clientes, mensagem=mensagem, erro=erro)

@app.route('/editar_fornecedor/<int:id>', methods=['GET', 'POST'])
def editar_fornecedor(id):
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT ClienteID, Nome, CPF FROM Clientes ORDER BY Nome")
    clientes = cursor.fetchall()
    if request.method == 'POST':
        cliente_id          = request.form.get('cliente_id')
        nome_fornecedor     = request.form.get('nome_fornecedor', '')
        email               = request.form.get('email', '')
        telefone            = request.form.get('telefone', '')
        nome_contato        = request.form.get('nome_contato', '')
        fornece_referencia  = request.form.get('fornece_referencia', 'Sim')
        limite_credito      = limpar_valor(request.form.get('limite_credito', 0))
        maior_compra_data   = request.form.get('maior_compra_data') or None
        maior_compra_valor  = limpar_valor(request.form.get('maior_compra_valor', 0))
        ultima_compra_data  = request.form.get('ultima_compra_data') or None
        ultima_compra_valor = limpar_valor(request.form.get('ultima_compra_valor', 0))
        data_cadastro       = request.form.get('data_cadastro_fornecedor') or None
        cursor.execute("""
            UPDATE Fornecedores SET
                ClienteID=%s, NomeFornecedor=%s, Email=%s, Telefone=%s,
                NomeContato=%s, ForneceReferencia=%s, LimiteCredito=%s,
                PrimeiraCompraData=%s, PrimeiraCompraValor=%s,
                UltimaCompraData=%s, UltimaCompraValor=%s,
                DataCadastroFornecedor=%s
            WHERE FornecedorID=%s
        """, (cliente_id, nome_fornecedor, email, telefone, nome_contato,
              fornece_referencia, limite_credito, maior_compra_data,
              maior_compra_valor, ultima_compra_data, ultima_compra_valor,
              data_cadastro, id))
        conn.commit()
        conn.close()
        return redirect(url_for('fornecedores'))
    cursor.execute("""
        SELECT FornecedorID, ClienteID, NomeFornecedor, Email, Telefone,
               NomeContato, ForneceReferencia, LimiteCredito,
               PrimeiraCompraData, PrimeiraCompraValor,
               UltimaCompraData, UltimaCompraValor, DataCadastroFornecedor
        FROM Fornecedores WHERE FornecedorID=%s
    """, (id,))
    fornecedor = cursor.fetchone()
    conn.close()
    return render_template('editar_fornecedor.html',
                           fornecedor=fornecedor, clientes=clientes)

@app.route('/excluir_fornecedor/<int:id>')
def excluir_fornecedor(id):
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Fornecedores WHERE FornecedorID=%s", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('fornecedores'))

# ─── NOTAS FISCAIS ────────────────────────────────────────────

@app.route('/notas_fiscais/<int:fornecedor_id>', methods=['GET', 'POST'])
def notas_fiscais(fornecedor_id):
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT F.NomeFornecedor, C.Nome
        FROM Fornecedores F
        INNER JOIN Clientes C ON F.ClienteID = C.ClienteID
        WHERE F.FornecedorID = %s
    """, (fornecedor_id,))
    info = cursor.fetchone()
    nome_fornecedor = info[0] if info else ''
    nome_cliente    = info[1] if info else ''
    mensagem = None
    if request.method == 'POST':
        numero_nf       = request.form.get('numero_nf', '')
        data_emissao    = request.form.get('data_emissao')
        valor_nf        = limpar_valor(request.form.get('valor_nf', 0))
        data_vencimento = request.form.get('data_vencimento')
        cursor.execute("""
            INSERT INTO NotasFiscais
            (FornecedorID, NumeroNF, DataEmissao, ValorNF, DataVencimento)
            VALUES (%s, %s, %s, %s, %s)
        """, (fornecedor_id, numero_nf, data_emissao, valor_nf, data_vencimento))
        conn.commit()
        mensagem = f'✅ NF {numero_nf} cadastrada com sucesso!'
    cursor.execute("""
        SELECT NFID, NumeroNF, DataEmissao, ValorNF, DataVencimento
        FROM NotasFiscais
        WHERE FornecedorID = %s
        ORDER BY DataVencimento DESC
    """, (fornecedor_id,))
    notas = cursor.fetchall()
    conn.close()
    return render_template('notas_fiscais.html',
                           notas=notas,
                           fornecedor_id=fornecedor_id,
                           nome_fornecedor=nome_fornecedor,
                           nome_cliente=nome_cliente,
                           mensagem=mensagem,
                           now=datetime.now())

@app.route('/excluir_nf/<int:nf_id>/<int:fornecedor_id>')
def excluir_nf(nf_id, fornecedor_id):
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM NotasFiscais WHERE NFID=%s", (nf_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('notas_fiscais', fornecedor_id=fornecedor_id))

# ─── DASHBOARD ────────────────────────────────────────────────

@app.route('/dashboard')
def dashboard():
    if login_necessario():
        return redirect(url_for('login'))
    conn   = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT ClienteID, Situacao FROM Clientes")
    clientes = cursor.fetchall()
    clientes_total = len(clientes)
    aprovados  = sum(1 for c in clientes if c[1] == 'Aprovado')
    pendentes  = sum(1 for c in clientes if c[1] == 'Pendente')
    recusados  = sum(1 for c in clientes if c[1] == 'Recusado')
    pct_aprovados = round((aprovados / clientes_total * 100)) if clientes_total > 0 else 0

    cursor.execute("""
        SELECT PontuacaoCredito FROM AnalisesCredito
        WHERE PontuacaoCredito IS NOT NULL
    """)
    scores = [r[0] for r in cursor.fetchall()]
    score_excelente = sum(1 for s in scores if s >= 851)
    score_bom       = sum(1 for s in scores if 701 <= s <= 850)
    score_regular   = sum(1 for s in scores if 501 <= s <= 700)
    score_baixo     = sum(1 for s in scores if s <= 500)

    cursor.execute("SELECT COUNT(*) FROM Fornecedores")
    total_fornecedores = cursor.fetchone()[0]

    # Lista de clientes para o filtro de fornecedores no dashboard
    cursor.execute("SELECT ClienteID, Nome FROM Clientes ORDER BY Nome")
    lista_clientes = cursor.fetchall()

    # Fornecedores do cliente selecionado (se houver filtro)
    cliente_selecionado_id   = request.args.get('cliente_id', type=int)
    cliente_selecionado_nome = None
    fornecedores_do_cliente  = []
    if cliente_selecionado_id:
        cursor.execute("SELECT Nome FROM Clientes WHERE ClienteID = %s", (cliente_selecionado_id,))
        row = cursor.fetchone()
        cliente_selecionado_nome = row[0] if row else None
        cursor.execute("""
            SELECT NomeFornecedor, Email, Telefone, NomeContato,
                   ForneceReferencia, LimiteCredito, UltimaCompraData, FornecedorID
            FROM Fornecedores
            WHERE ClienteID = %s
            ORDER BY NomeFornecedor
        """, (cliente_selecionado_id,))
        fornecedores_do_cliente = cursor.fetchall()

    conn.close()

    return render_template('dashboard.html',
        clientes_total=clientes_total,
        aprovados=aprovados,
        pendentes=pendentes,
        recusados=recusados,
        pct_aprovados=pct_aprovados,
        score_excelente=score_excelente,
        score_bom=score_bom,
        score_regular=score_regular,
        score_baixo=score_baixo,
        total_fornecedores=total_fornecedores,
        lista_clientes=lista_clientes,
        cliente_selecionado_id=cliente_selecionado_id,
        cliente_selecionado_nome=cliente_selecionado_nome,
        fornecedores_do_cliente=fornecedores_do_cliente)

# ─── RELATÓRIOS ───────────────────────────────────────────────

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from flask import make_response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

# ── Relatório Clientes PDF (todos ou por cliente) ──

@app.route('/relatorio/pdf')
def relatorio_pdf():
    if login_necessario():
        return redirect(url_for('login'))

    cliente_id = request.args.get('cliente_id', type=int)

    conn   = conectar()
    cursor = conn.cursor()

    if cliente_id:
        cursor.execute("""
            SELECT C.Nome, C.CPF, C.Situacao, C.HistoricoPagamento,
                   C.DividasAtuais, C.SetorAtuacao, C.SerasaObservacao,
                   AC.PontuacaoCredito, AC.LimiteSugerido, AC.TaxaJuros,
                   AC.DataAnalise
            FROM Clientes C
            LEFT JOIN AnalisesCredito AC ON AC.ClienteID = C.ClienteID
            WHERE C.ClienteID = %s
            ORDER BY C.Nome
        """, (cliente_id,))
    else:
        cursor.execute("""
            SELECT C.Nome, C.CPF, C.Situacao, C.HistoricoPagamento,
                   C.DividasAtuais, C.SetorAtuacao, C.SerasaObservacao,
                   AC.PontuacaoCredito, AC.LimiteSugerido, AC.TaxaJuros,
                   AC.DataAnalise
            FROM Clientes C
            LEFT JOIN AnalisesCredito AC ON AC.ClienteID = C.ClienteID
            ORDER BY C.Nome
        """)
    dados = cursor.fetchall()

    # Lista de clientes para o template de relatórios
    cursor.execute("SELECT ClienteID, Nome FROM Clientes ORDER BY Nome")
    lista_clientes = cursor.fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               rightMargin=1.5*cm, leftMargin=1.5*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles    = getSampleStyleSheet()
    elementos = []

    titulo_style = ParagraphStyle('titulo', fontSize=16, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1a2332'), spaceAfter=6)
    sub_style    = ParagraphStyle('sub', fontSize=10, fontName='Helvetica',
                                  textColor=colors.grey, spaceAfter=20)

    titulo_txt = f'Relatório Individual — {dados[0][0]}' if (cliente_id and dados) else 'Relatório de Clientes'
    elementos.append(Paragraph('Sistema de Análise de Crédito', titulo_style))
    elementos.append(Paragraph(f'{titulo_txt} — gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}', sub_style))

    cabecalho = ['Cliente', 'CPF/CNPJ', 'Score', 'Situação', 'Limite', 'Taxa', 'Histórico', 'Setor']
    linhas    = [cabecalho]

    for d in dados:
        situacao = d[2] or 'Pendente'
        linhas.append([
            Paragraph(str(d[0] or ''), styles['Normal']),
            str(formatar_documento(d[1]) or ''),
            str(d[7] or '—'),
            str(situacao),
            str(formatar_moeda(d[8])),
            f"{d[9]}% a.m." if d[9] else '—',
            str(d[3] or ''),
            str(d[5] or ''),
        ])

    tabela = Table(linhas, colWidths=[4.5*cm, 3.8*cm, 1.5*cm, 2.2*cm, 2.8*cm, 2*cm, 2*cm, 2*cm])
    tabela.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#1a2332')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 9),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f4f6fb')]),
        ('FONTSIZE',    (0,1), (-1,-1), 8),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('TOPPADDING',  (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
    ]))

    elementos.append(tabela)
    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph(f'Total de registros: {len(dados)}', styles['Normal']))

    doc.build(elementos)
    buffer.seek(0)

    nome_arquivo = f'relatorio_{dados[0][0]}.pdf' if (cliente_id and dados) else 'relatorio_clientes.pdf'
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={nome_arquivo}'
    return response

# ── Relatório Clientes Excel (todos ou por cliente) ──

@app.route('/relatorio/excel')
def relatorio_excel():
    if login_necessario():
        return redirect(url_for('login'))

    cliente_id = request.args.get('cliente_id', type=int)

    conn   = conectar()
    cursor = conn.cursor()

    if cliente_id:
        cursor.execute("""
            SELECT C.Nome, C.CPF, C.Situacao, C.HistoricoPagamento,
                   C.DividasAtuais, C.SetorAtuacao, C.SerasaObservacao,
                   AC.PontuacaoCredito, AC.LimiteSugerido, AC.TaxaJuros,
                   AC.DataAnalise
            FROM Clientes C
            LEFT JOIN AnalisesCredito AC ON AC.ClienteID = C.ClienteID
            WHERE C.ClienteID = %s
            ORDER BY C.Nome
        """, (cliente_id,))
    else:
        cursor.execute("""
            SELECT C.Nome, C.CPF, C.Situacao, C.HistoricoPagamento,
                   C.DividasAtuais, C.SetorAtuacao, C.SerasaObservacao,
                   AC.PontuacaoCredito, AC.LimiteSugerido, AC.TaxaJuros,
                   AC.DataAnalise
            FROM Clientes C
            LEFT JOIN AnalisesCredito AC ON AC.ClienteID = C.ClienteID
            ORDER BY C.Nome
        """)
    dados = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Análise de Clientes'

    header_fill = PatternFill('solid', fgColor='1a2332')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border      = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb')
    )

    titulo_txt = f'Análise de Crédito — {dados[0][0]}' if (cliente_id and dados) else 'Sistema de Análise de Crédito — Relatório de Clientes'
    ws.merge_cells('A1:K1')
    ws['A1'] = titulo_txt
    ws['A1'].font      = Font(bold=True, size=14, color='1a2332')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:K2')
    ws['A2'] = f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font      = Font(size=10, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    cabecalhos = ['Nome', 'CPF/CNPJ', 'Situação', 'Histórico Pgto',
                  'Dívidas', 'Setor', 'Observação Serasa',
                  'Score', 'Limite', 'Taxa', 'Data Análise']
    for col, cab in enumerate(cabecalhos, 1):
        cell           = ws.cell(row=4, column=col, value=cab)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border    = border

    alt_fill = PatternFill('solid', fgColor='f4f6fb')
    for row_idx, d in enumerate(dados, 5):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        linha = [
            d[0] or '',
            formatar_documento(d[1]) or '',
            d[2] or 'Pendente',
            d[3] or '',
            float(d[4] or 0),
            d[5] or '',
            d[6] or '',
            d[7] or '',
            float(d[8] or 0),
            f"{d[9]}% a.m." if d[9] else '—',
            d[10].strftime('%d/%m/%Y %H:%M') if d[10] else '—',
        ]
        for col_idx, valor in enumerate(linha, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(horizontal='center')

    larguras = [35, 22, 12, 16, 14, 18, 30, 8, 16, 12, 20]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f'relatorio_{dados[0][0]}.xlsx' if (cliente_id and dados) else 'relatorio_clientes.xlsx'
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={nome_arquivo}'
    return response

# ── Relatório Fornecedores PDF (todos ou por cliente) ──

@app.route('/relatorio/fornecedores/pdf')
def relatorio_fornecedores_pdf():
    if login_necessario():
        return redirect(url_for('login'))

    cliente_id = request.args.get('cliente_id', type=int)

    conn   = conectar()
    cursor = conn.cursor()

    if cliente_id:
        cursor.execute("""
            SELECT C.Nome, C.CPF, F.NomeFornecedor, F.Email, F.Telefone,
                   F.NomeContato, F.ForneceReferencia, F.LimiteCredito,
                   F.PrimeiraCompraData, F.PrimeiraCompraValor,
                   F.UltimaCompraData, F.UltimaCompraValor,
                   F.DataCadastroFornecedor, F.FornecedorID
            FROM Fornecedores F
            INNER JOIN Clientes C ON F.ClienteID = C.ClienteID
            WHERE C.ClienteID = %s
            ORDER BY F.NomeFornecedor
        """, (cliente_id,))
    else:
        cursor.execute("""
            SELECT C.Nome, C.CPF, F.NomeFornecedor, F.Email, F.Telefone,
                   F.NomeContato, F.ForneceReferencia, F.LimiteCredito,
                   F.PrimeiraCompraData, F.PrimeiraCompraValor,
                   F.UltimaCompraData, F.UltimaCompraValor,
                   F.DataCadastroFornecedor, F.FornecedorID
            FROM Fornecedores F
            INNER JOIN Clientes C ON F.ClienteID = C.ClienteID
            ORDER BY C.Nome, F.NomeFornecedor
        """)
    fornecedores = cursor.fetchall()

    nfs_por_fornecedor = {}
    if fornecedores:
        ids = tuple(set(f[13] for f in fornecedores))
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            SELECT FornecedorID, NumeroNF, DataEmissao, ValorNF, DataVencimento
            FROM NotasFiscais
            WHERE FornecedorID IN ({placeholders})
            ORDER BY FornecedorID, DataVencimento DESC
        """, ids)
        for nf in cursor.fetchall():
            fid = nf[0]
            if fid not in nfs_por_fornecedor:
                nfs_por_fornecedor[fid] = []
            nfs_por_fornecedor[fid].append(nf)
    conn.close()

    buffer  = io.BytesIO()
    doc     = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
    styles  = getSampleStyleSheet()
    elementos = []

    titulo_style  = ParagraphStyle('titulo', fontSize=16, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#1a2332'), spaceAfter=4)
    sub_style     = ParagraphStyle('sub', fontSize=10, fontName='Helvetica',
                                   textColor=colors.grey, spaceAfter=16)
    cliente_style = ParagraphStyle('cliente', fontSize=12, fontName='Helvetica-Bold',
                                   textColor=colors.white, spaceAfter=4)
    forn_style    = ParagraphStyle('forn', fontSize=10, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#1a2332'), spaceAfter=4)

    titulo_txt = f'Fornecedores — {fornecedores[0][0]}' if (cliente_id and fornecedores) else 'Relatório de Fornecedores'
    elementos.append(Paragraph(f'Sistema de Análise de Crédito — {titulo_txt}', titulo_style))
    elementos.append(Paragraph(f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}', sub_style))

    cliente_atual = None
    for f in fornecedores:
        if f[0] != cliente_atual:
            cliente_atual = f[0]
            elementos.append(Spacer(1, 10))
            cabecalho_cliente = Table(
                [[Paragraph(f'👤 {f[0]}  —  {formatar_documento(f[1])}', cliente_style)]],
                colWidths=[18*cm]
            )
            cabecalho_cliente.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#1a2332')),
                ('TOPPADDING',    (0,0), (-1,-1), 8),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ('LEFTPADDING',   (0,0), (-1,-1), 12),
            ]))
            elementos.append(cabecalho_cliente)
            elementos.append(Spacer(1, 6))

        elementos.append(Paragraph(f'🏭 {f[2]}', forn_style))
        dados_forn = [
            ['Email', f[3] or '—', 'Telefone', f[4] or '—'],
            ['Contato', f[5] or '—', 'Dá Referência', f[6] or '—'],
            ['Limite de Crédito', formatar_moeda(f[7]), 'Cad. Fornecedor',
             f[12].strftime('%d/%m/%Y') if f[12] else '—'],
            ['Maior Compra',
             f"{f[8].strftime('%d/%m/%Y') if f[8] else '—'} — {formatar_moeda(f[9])}",
             'Última Compra',
             f"{f[10].strftime('%d/%m/%Y') if f[10] else '—'} — {formatar_moeda(f[11])}"],
        ]
        tabela_forn = Table(dados_forn, colWidths=[3.5*cm, 5.5*cm, 3.5*cm, 5.5*cm])
        tabela_forn.setStyle(TableStyle([
            ('FONTSIZE',    (0,0), (-1,-1), 8),
            ('FONTNAME',    (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME',    (2,0), (2,-1), 'Helvetica-Bold'),
            ('TEXTCOLOR',   (0,0), (0,-1), colors.HexColor('#64748b')),
            ('TEXTCOLOR',   (2,0), (2,-1), colors.HexColor('#64748b')),
            ('BACKGROUND',  (0,0), (-1,-1), colors.HexColor('#f8fafc')),
            ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
            ('TOPPADDING',  (0,0), (-1,-1), 5),
            ('BOTTOMPADDING',(0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        elementos.append(tabela_forn)

        fid = f[13]
        nfs = nfs_por_fornecedor.get(fid, [])
        if nfs:
            elementos.append(Spacer(1, 4))
            nf_cab = [['Chave de Acesso', 'Emissão', 'Valor', 'Vencimento']]
            hoje   = datetime.now().date()
            for nf in nfs:
                venc     = nf[4]
                venc_str = venc.strftime('%d/%m/%Y') if venc else '—'
                if venc and venc < hoje:
                    venc_str += ' ⚠ Vencida'
                nf_cab.append([
                    str(nf[1]),
                    nf[2].strftime('%d/%m/%Y') if nf[2] else '—',
                    formatar_moeda(nf[3]),
                    venc_str
                ])
            tabela_nf = Table(nf_cab, colWidths=[7*cm, 2.5*cm, 3*cm, 2.5*cm])
            tabela_nf.setStyle(TableStyle([
                ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#4f9cf9')),
                ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
                ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',    (0,0), (-1,-1), 8),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eff6ff')]),
                ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
                ('TOPPADDING',  (0,0), (-1,-1), 4),
                ('BOTTOMPADDING',(0,0), (-1,-1), 4),
                ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
            ]))
            elementos.append(tabela_nf)
        else:
            elementos.append(Spacer(1, 2))
            elementos.append(Paragraph('Nenhuma NF cadastrada.',
                             ParagraphStyle('sem_nf', fontSize=8, textColor=colors.grey)))

        elementos.append(Spacer(1, 12))

    doc.build(elementos)
    buffer.seek(0)
    nome_arquivo = f'fornecedores_{fornecedores[0][0]}.pdf' if (cliente_id and fornecedores) else 'relatorio_fornecedores.pdf'
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={nome_arquivo}'
    return response

# ── Relatório Fornecedores Excel (todos ou por cliente) ──

@app.route('/relatorio/fornecedores/excel')
def relatorio_fornecedores_excel():
    if login_necessario():
        return redirect(url_for('login'))

    cliente_id = request.args.get('cliente_id', type=int)

    conn   = conectar()
    cursor = conn.cursor()

    if cliente_id:
        cursor.execute("""
            SELECT C.Nome, C.CPF, F.NomeFornecedor, F.Email, F.Telefone,
                   F.NomeContato, F.ForneceReferencia, F.LimiteCredito,
                   F.PrimeiraCompraData, F.PrimeiraCompraValor,
                   F.UltimaCompraData, F.UltimaCompraValor,
                   F.DataCadastroFornecedor, F.FornecedorID
            FROM Fornecedores F
            INNER JOIN Clientes C ON F.ClienteID = C.ClienteID
            WHERE C.ClienteID = %s
            ORDER BY F.NomeFornecedor
        """, (cliente_id,))
    else:
        cursor.execute("""
            SELECT C.Nome, C.CPF, F.NomeFornecedor, F.Email, F.Telefone,
                   F.NomeContato, F.ForneceReferencia, F.LimiteCredito,
                   F.PrimeiraCompraData, F.PrimeiraCompraValor,
                   F.UltimaCompraData, F.UltimaCompraValor,
                   F.DataCadastroFornecedor, F.FornecedorID
            FROM Fornecedores F
            INNER JOIN Clientes C ON F.ClienteID = C.ClienteID
            ORDER BY C.Nome, F.NomeFornecedor
        """)
    fornecedores = cursor.fetchall()

    nfs_por_fornecedor = {}
    if fornecedores:
        ids = tuple(set(f[13] for f in fornecedores))
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f"""
            SELECT FornecedorID, NumeroNF, DataEmissao, ValorNF, DataVencimento
            FROM NotasFiscais
            WHERE FornecedorID IN ({placeholders})
            ORDER BY FornecedorID, DataVencimento DESC
        """, ids)
        for nf in cursor.fetchall():
            fid = nf[0]
            if fid not in nfs_por_fornecedor:
                nfs_por_fornecedor[fid] = []
            nfs_por_fornecedor[fid].append(nf)
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fornecedores'

    header_fill = PatternFill('solid', fgColor='1a2332')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill    = PatternFill('solid', fgColor='f4f6fb')
    border      = Border(
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb'),
        top=Side(style='thin', color='e5e7eb'),
        bottom=Side(style='thin', color='e5e7eb')
    )

    titulo_txt = f'Fornecedores — {fornecedores[0][0]}' if (cliente_id and fornecedores) else 'Sistema de Análise de Crédito — Relatório de Fornecedores'
    ws.merge_cells('A1:M1')
    ws['A1'] = titulo_txt
    ws['A1'].font      = Font(bold=True, size=14, color='1a2332')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font      = Font(size=10, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    cabs = ['Cliente', 'CPF/CNPJ', 'Fornecedor', 'Email', 'Telefone',
            'Contato', 'Referência', 'Limite', 'Data Maior Compra',
            'Valor Maior Compra', 'Data Última Compra', 'Valor Última Compra',
            'Cad. Fornecedor']
    for col, cab in enumerate(cabs, 1):
        cell           = ws.cell(row=4, column=col, value=cab)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border    = border

    for row_idx, f in enumerate(fornecedores, 5):
        fill  = alt_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        linha = [
            f[0], formatar_documento(f[1]), f[2], f[3] or '', f[4] or '',
            f[5] or '', f[6] or '', float(f[7] or 0),
            f[8].strftime('%d/%m/%Y') if f[8] else '—',
            float(f[9] or 0),
            f[10].strftime('%d/%m/%Y') if f[10] else '—',
            float(f[11] or 0),
            f[12].strftime('%d/%m/%Y') if f[12] else '—',
        ]
        for col_idx, valor in enumerate(linha, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(horizontal='center')

    larguras = [30, 22, 25, 28, 18, 20, 12, 16, 18, 18, 18, 18, 18]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = larg

    ws2 = wb.create_sheet('Notas Fiscais')
    ws2.merge_cells('A1:F1')
    ws2['A1'] = 'Notas Fiscais por Fornecedor'
    ws2['A1'].font      = Font(bold=True, size=14, color='1a2332')
    ws2['A1'].alignment = Alignment(horizontal='center')

    header_fill2 = PatternFill('solid', fgColor='4f9cf9')
    cabs2 = ['Cliente', 'Fornecedor', 'Chave de Acesso', 'Data Emissão', 'Valor', 'Vencimento']
    for col, cab in enumerate(cabs2, 1):
        cell           = ws2.cell(row=3, column=col, value=cab)
        cell.fill      = header_fill2
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center')
        cell.border    = border

    row_idx = 4
    hoje    = datetime.now().date()
    for f in fornecedores:
        fid = f[13]
        nfs = nfs_por_fornecedor.get(fid, [])
        for nf in nfs:
            fill    = alt_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            venc    = nf[4]
            vencida = ' ⚠ VENCIDA' if venc and venc < hoje else ''
            linha   = [
                f[0], f[2], str(nf[1]),
                nf[2].strftime('%d/%m/%Y') if nf[2] else '—',
                float(nf[3] or 0),
                (venc.strftime('%d/%m/%Y') if venc else '—') + vencida
            ]
            for col_idx, valor in enumerate(linha, 1):
                cell           = ws2.cell(row=row_idx, column=col_idx, value=valor)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = Alignment(horizontal='center')
                if vencida and col_idx == 6:
                    cell.font = Font(color='dc2626', bold=True)
            row_idx += 1

    for col, larg in enumerate([30, 25, 12, 16, 16, 22], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = larg

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome_arquivo = f'fornecedores_{fornecedores[0][0]}.xlsx' if (cliente_id and fornecedores) else 'relatorio_fornecedores.xlsx'
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={nome_arquivo}'
    return response

if __name__ == '__main__':
    app.run(debug=True)
